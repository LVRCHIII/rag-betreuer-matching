"""Erzeugt aus den Roh-Ergebnissen den Excel-Report, ein Markdown-Summary und JSON."""
import json
import statistics
from datetime import datetime
from typing import List, Dict, Any

# Zielwerte (KPIs) – entsprechen der Tabelle in Kapitel 6.5 der Projektdokumentation
TARGETS = {
    "top3": 0.70,
    "faithfulness": 0.85,
    "answer_relevancy": 0.80,
    "context_recall": 0.75,
}


def _mean(xs):
    xs = [x for x in xs if x is not None]
    return round(statistics.mean(xs), 4) if xs else None


def _median(xs):
    xs = [x for x in xs if x is not None]
    return round(statistics.median(xs), 4) if xs else None


def _rate(flags):
    flags = [bool(f) for f in flags if f is not None]
    return round(sum(flags) / len(flags), 4) if flags else None


def aggregate(records: List[Dict[str, Any]], meta: Dict[str, Any]) -> Dict[str, Any]:
    A = [r for r in records if r["layer"] in ("A", "B") and not r.get("error")]
    B = [r for r in records if r["layer"] == "B" and not r.get("error")]
    C = [r for r in records if r["layer"] == "C" and not r.get("error")]

    agg = {
        "meta": meta,
        "counts": {"matching": len(A), "ragas": len(B), "robustness": len(C),
                   "errors": sum(1 for r in records if r.get("error"))},
        "layer_a": {
            "hit@1": _rate([r.get("hit@1") for r in A]),
            "hit@3": _rate([r.get("hit@3") for r in A]),
            "hit@5": _rate([r.get("hit@5") for r in A]),
            "mrr": _mean([r.get("reciprocal_rank") for r in A]),
        },
        "layer_b": {
            "faithfulness_mean": _mean([r.get("faithfulness") for r in B]),
            "faithfulness_median": _median([r.get("faithfulness") for r in B]),
            "answer_relevancy_mean": _mean([r.get("answer_relevancy") for r in B]),
            "answer_relevancy_median": _median([r.get("answer_relevancy") for r in B]),
            "context_recall_mean": _mean([r.get("context_recall") for r in B]),
            "overall_mean": _mean([r.get("overall") for r in B]),
        },
        "layer_c": {
            "correct_refusal": _rate([r.get("correct_refusal") for r in C]),
            "fallback_ok": _rate([r.get("fallback_ok") for r in C]),
            "declined": _rate([r.get("declined") for r in C]),
            "german": _rate([r.get("german") for r in C]),
        },
    }
    return agg


def _kpi_rows(agg: Dict[str, Any]):
    a, b = agg["layer_a"], agg["layer_b"]
    def status(val, tgt):
        if val is None:
            return "—"
        return "erfüllt" if val >= tgt else "nicht erfüllt"
    return [
        ("Matching-Relevanz – Top-3-Trefferquote", a["hit@3"], TARGETS["top3"], status(a["hit@3"], TARGETS["top3"])),
        ("Quelltreue – Faithfulness (RAGAS)", b["faithfulness_mean"], TARGETS["faithfulness"], status(b["faithfulness_mean"], TARGETS["faithfulness"])),
        ("Antwortrelevanz – Answer Relevancy (RAGAS)", b["answer_relevancy_mean"], TARGETS["answer_relevancy"], status(b["answer_relevancy_mean"], TARGETS["answer_relevancy"])),
        ("Kontextabdeckung – Context Recall (RAGAS)", b["context_recall_mean"], TARGETS["context_recall"], status(b["context_recall_mean"], TARGETS["context_recall"])),
    ]


def write_excel(records, agg, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="132933")
    head_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor="D7F0DE")
    bad_fill = PatternFill("solid", fgColor="F6D9D2")

    def header(ws, row, cols, widths=None):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=row, column=i, value=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center")
            if widths:
                ws.column_dimensions[cell.column_letter].width = widths[i - 1]

    # --- Blatt 1: Zusammenfassung ---
    ws = wb.active
    ws.title = "Zusammenfassung"
    ws["A1"] = "RAG Betreuer-Matching – Evaluations-Report"
    ws["A1"].font = Font(bold=True, size=14, color="132933")
    m = agg["meta"]
    info = [
        ("Datum", m.get("date")), ("Sprachmodell (LLM)", m.get("model")),
        ("Embedding-Modell", m.get("embedding_model")), ("Bewerter (RAGAS-Judge)", m.get("judge_model")),
        ("Collection", m.get("collection")), ("Distraktor-Collections", m.get("extra_collections", "-")),
        ("Top-k", m.get("k")), ("Min-Score (Kontext)", m.get("min_score", 0.0)),
        ("System-Prompt-Variante", m.get("prompt_variant", "current")),
        ("Git-Commit", m.get("commit")),
        ("Fragen gesamt", agg["counts"]["matching"] + agg["counts"]["robustness"]),
        ("davon RAGAS-bewertet", agg["counts"]["ragas"]),
        ("Robustheits-Fragen", agg["counts"]["robustness"]),
        ("Fehler", agg["counts"]["errors"]),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48

    r += 1
    ws.cell(row=r, column=1, value="KPI-Abgleich").font = Font(bold=True, size=12, color="132933")
    r += 1
    header(ws, r, ["Kennzahl", "Ergebnis", "Zielwert", "Status"], [44, 14, 12, 16])
    r += 1
    for name, val, tgt, st in _kpi_rows(agg):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val if val is not None else "—")
        ws.cell(row=r, column=3, value=f"≥ {tgt}")
        c = ws.cell(row=r, column=4, value=st)
        c.fill = ok_fill if st == "erfüllt" else (bad_fill if st == "nicht erfüllt" else PatternFill())
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Weitere Kennzahlen").font = Font(bold=True, size=12, color="132933")
    r += 1
    extra = [
        ("Top-1-Trefferquote", agg["layer_a"]["hit@1"]),
        ("Top-5-Trefferquote", agg["layer_a"]["hit@5"]),
        ("Mean Reciprocal Rank (MRR)", agg["layer_a"]["mrr"]),
        ("Robustheit – korrekte Ablehnung (LLM-Judge, Ebene C)", agg["layer_c"].get("correct_refusal")),
        ("Sprach-Compliance Deutsch (Ebene C)", agg["layer_c"]["german"]),
    ]
    for k, v in extra:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v if v is not None else "—")
        r += 1

    # --- Blatt 2: Pro Frage ---
    ws2 = wb.create_sheet("Pro Frage")
    cols = ["ID", "Ebene", "Art", "Frage", "Erwartet", "FB", "Rang",
            "Hit@1", "Hit@3", "Hit@5", "Faithfulness", "Answer Rel.", "Context Recall",
            "Fallback ok", "Deutsch", "Antwort (gekürzt)"]
    header(ws2, 1, cols, [8, 7, 10, 46, 22, 6, 6, 7, 7, 7, 12, 12, 13, 10, 8, 60])
    for i, rec in enumerate(records, 2):
        ws2.cell(row=i, column=1, value=rec.get("id"))
        ws2.cell(row=i, column=2, value=rec.get("layer"))
        ws2.cell(row=i, column=3, value=rec.get("kind"))
        ws2.cell(row=i, column=4, value=rec.get("question"))
        ws2.cell(row=i, column=5, value=rec.get("expected"))
        ws2.cell(row=i, column=6, value=rec.get("fachbereich"))
        ws2.cell(row=i, column=7, value=rec.get("rank"))
        ws2.cell(row=i, column=8, value=rec.get("hit@1"))
        ws2.cell(row=i, column=9, value=rec.get("hit@3"))
        ws2.cell(row=i, column=10, value=rec.get("hit@5"))
        ws2.cell(row=i, column=11, value=rec.get("faithfulness"))
        ws2.cell(row=i, column=12, value=rec.get("answer_relevancy"))
        ws2.cell(row=i, column=13, value=rec.get("context_recall"))
        ws2.cell(row=i, column=14, value=rec.get("fallback_ok"))
        ws2.cell(row=i, column=15, value=rec.get("german"))
        ans = (rec.get("answer") or "")[:300]
        ws2.cell(row=i, column=16, value=ans)
    ws2.freeze_panes = "A2"

    # --- Blatt 3: Verteilungen ---
    ws3 = wb.create_sheet("Verteilungen")
    bins = ["0.0–0.5", "0.5–0.7", "0.7–0.85", "0.85–1.0"]

    def histo(values):
        c = [0, 0, 0, 0]
        for v in values:
            if v is None:
                continue
            if v < 0.5: c[0] += 1
            elif v < 0.7: c[1] += 1
            elif v < 0.85: c[2] += 1
            else: c[3] += 1
        return c

    B = [r for r in records if r["layer"] == "B" and not r.get("error")]
    header(ws3, 1, ["Score-Bereich", "Faithfulness", "Answer Relevancy", "Context Recall"], [14, 16, 18, 16])
    f_h = histo([r.get("faithfulness") for r in B])
    a_h = histo([r.get("answer_relevancy") for r in B])
    c_h = histo([r.get("context_recall") for r in B])
    for i, b in enumerate(bins):
        ws3.cell(row=2 + i, column=1, value=b)
        ws3.cell(row=2 + i, column=2, value=f_h[i])
        ws3.cell(row=2 + i, column=3, value=a_h[i])
        ws3.cell(row=2 + i, column=4, value=c_h[i])
    chart = BarChart()
    chart.title = "Verteilung der RAGAS-Scores (Ebene B)"
    chart.y_axis.title = "Anzahl Antworten"
    chart.x_axis.title = "Score-Bereich"
    data = Reference(ws3, min_col=2, max_col=4, min_row=1, max_row=5)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws3.add_chart(chart, "F2")

    wb.save(path)


def write_markdown(agg, path):
    a, b, c = agg["layer_a"], agg["layer_b"], agg["layer_c"]
    m = agg["meta"]

    def pct(x):
        return f"{x*100:.1f} %" if x is not None else "—"

    def num(x):
        return f"{x:.3f}" if x is not None else "—"

    lines = [
        "# Evaluations-Report – RAG Betreuer-Matching", "",
        f"- **Datum:** {m.get('date')}",
        f"- **Sprachmodell:** {m.get('model')} · **Embedding:** {m.get('embedding_model')} · **RAGAS-Judge:** {m.get('judge_model')}",
        f"- **Collection:** {m.get('collection')} (+Distraktoren: {m.get('extra_collections', '-')}) · "
        f"**Top-k:** {m.get('k')} · **Min-Score:** {m.get('min_score', 0.0)} · "
        f"**Prompt:** {m.get('prompt_variant', 'current')} · **Commit:** {m.get('commit')}",
        f"- **Fragen:** {agg['counts']['matching'] + agg['counts']['robustness']} gesamt, "
        f"davon {agg['counts']['ragas']} RAGAS-bewertet, {agg['counts']['robustness']} Robustheit", "",
        "## KPI-Abgleich", "",
        "| Kennzahl | Ergebnis | Zielwert | Status |",
        "|---|---|---|---|",
    ]
    for name, val, tgt, st in _kpi_rows(agg):
        shown = pct(val) if "Trefferquote" in name else num(val)
        lines.append(f"| {name} | {shown} | ≥ {tgt} | {st} |")
    lines += [
        "", "## Ebene A – Matching-Genauigkeit", "",
        f"- Top-1: **{pct(a['hit@1'])}** · Top-3: **{pct(a['hit@3'])}** · Top-5: **{pct(a['hit@5'])}**",
        f"- Mean Reciprocal Rank: **{num(a['mrr'])}**", "",
        "## Ebene B – Antwortqualität (RAGAS)", "",
        f"- Faithfulness: **{num(b['faithfulness_mean'])}** (Median {num(b['faithfulness_median'])})",
        f"- Answer Relevancy: **{num(b['answer_relevancy_mean'])}** (Median {num(b['answer_relevancy_median'])})",
        f"- Context Recall: **{num(b['context_recall_mean'])}**", "",
        "## Ebene C – Robustheit & Compliance", "",
        f"- Korrekte Ablehnung unpassender Anfragen (LLM-Judge): **{pct(c.get('correct_refusal'))}**",
        f"- (heuristisch, nachrichtlich): {pct(c.get('fallback_ok'))}",
        f"- Sprach-Compliance (Deutsch): **{pct(c['german'])}**", "",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_all(records, meta, out_dir):
    import os
    os.makedirs(out_dir, exist_ok=True)
    agg = aggregate(records, meta)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = os.path.join(out_dir, f"eval_report_{stamp}.xlsx")
    md = os.path.join(out_dir, f"eval_summary_{stamp}.md")
    js = os.path.join(out_dir, f"eval_raw_{stamp}.json")
    write_excel(records, agg, xlsx)
    write_markdown(agg, md)
    with open(js, "w", encoding="utf-8") as f:
        json.dump({"aggregate": agg, "records": records}, f, ensure_ascii=False, indent=2)
    return xlsx, md, js
